#_______________________system libraries
import os
cwd = os.getcwd()
print(cwd)
#import random
import time
#import pyautogui
import openpyxl  as O
#import pydub
import os
#import xlrd
#import ffmpy
excel_file= "LIST.xlsx"
Excel_worksheet="Sheet1"
wb=O.load_workbook(excel_file)
ws=wb[Excel_worksheet]
row_num=ws.max_row
col_num=ws.max_column
print("the mo. of rows is " ,row_num ,"and the number of columns is " , col_num)
row = 3
print("ID = " ,ws.cell(row ,1).value)
print("code_meli = " ,ws.cell(row ,1).value)
